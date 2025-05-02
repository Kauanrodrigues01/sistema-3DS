import csv
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from django.http import HttpResponse
from .models import Selection

def download_csv_selections_report(request):
    """
    Gera e faz o download do relatório CSV de todas as seleções.
    """
    selections = Selection.objects.select_related('teacher')

    # Configura a resposta HTTP para download do CSV
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="selections_report.csv"'

    writer = csv.writer(response)

    # Escreve o cabeçalho do CSV
    writer.writerow(["Nome do Aluno", "Professor (Horário)"])

    # Escreve os dados das seleções
    for selection in selections:
        writer.writerow([
            selection.student_name,
            selection.teacher.name
        ])

    return response


def download_excel_selections_report(request):
    """
    Gera e faz o download do relatório Excel de todas as seleções.
    """
    selections = Selection.objects.select_related('teacher').order_by('teacher__name')

    # Criar um arquivo Excel
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Relatório de Seleções"

    # Estilização do cabeçalho
    headers = ["Nome do Aluno", "Professor (Horário)"]
    sheet.append(headers)

    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    header_alignment = Alignment(horizontal="center", vertical="center")

    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    # Estilização das linhas (cor alternada e bordas)
    row_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    for row_index, selection in enumerate(selections, start=2):
        row_data = [
            selection.student_name,
            selection.teacher.name
        ]
        sheet.append(row_data)

        for cell in sheet[row_index]:
            cell.border = thin_border
            if row_index % 2 == 0:
                cell.fill = row_fill

    # Ajustar automaticamente a largura das colunas
    for col in sheet.columns:
        max_length = 0
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        sheet.column_dimensions[col_letter].width = max_length + 5

    # Configurar a resposta HTTP para download do Excel
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="selections_report.xlsx"'

    workbook.save(response)

    return response
